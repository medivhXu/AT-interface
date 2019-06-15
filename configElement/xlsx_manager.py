# !/uer/bin/env python3

import datetime
import os

try:
    from openpyxl import load_workbook
except ImportError:
    os.system("pip3 install openpyxl")
    from openpyxl import load_workbook

from openpyxl.styles import Font, colors, Alignment


class ConfExcel(object):
    def __init__(self, fp, sheet_name="用例", case_start_row='A8'):
        self._wb = load_workbook(fp)
        self._sheet = self._wb[sheet_name]
        self._case = []
        self.case_start_row = case_start_row

    def get_email_data(self):
        """取excel中配置的email数据"""
        smtp_server = self._sheet['G3'].value
        receiver = self._sheet['G4'].value
        sender = self._sheet['G5'].value
        sender_pwd = self._sheet['G6'].value
        self._wb.close()
        return smtp_server, receiver, sender, sender_pwd

    def get_public_data(self):
        """去接口参数的公共部分"""
        app_secret = self._sheet['F3'].value
        app_key = self._sheet['H3'].value
        version = self._sheet['B2'].value
        token = self._sheet['F5'].value
        return app_secret, app_key, version, token

    def get_cases_data(self):
        """去所有接口，返回接口列表"""
        index_no = int(self.case_start_row[1]) - 1
        while index_no <= self._sheet.max_row - 1:
            case_dict = {column[6].value: column[index_no].value for column in self._sheet.columns}
            self._case.append(case_dict)
            index_no += 1
        self._wb.close()
        return self._case

    def write_the_result_to_the_new_excel(self, result, differences=None, file_fp=None, report_dir_name='report'):
        """
        把结果写到excel里
        :param result:
            示例:
                result = {case_01: Ture, case_02: False}
        :param differences:
        :param file_fp:
        :param report_dir_name:
        :return:
        """
        end_column = self._sheet.max_column - 1
        no = 0
        for i in range(int(self.case_start_row[1]), self._sheet.max_row + 1):
            if differences is not None:
                self._sheet.cell(row=i, column=end_column).value = result[no] or '不通过'
                # 设置不通过字体和颜色
                bold_24_font = Font(size=24, color=colors.RED, bold=True)
                self._sheet.cell(row=i, column=end_column).font = bold_24_font
                self._sheet.cell(row=i, column=end_column).alignment = Alignment(horizontal='center', vertical='center')
                self._sheet.cell(row=i, column=self._sheet.max_column).value = differences[no]
                no += 1
        if file_fp is None:
            now = datetime.datetime.now().strftime('%Y-%m-%d_%H_%M_%S')
            __dir__ = os.path.dirname(os.path.abspath(__file__))
            # 如果没有报告目录就创建一个
            if report_dir_name not in os.listdir(os.path.join(__dir__, '../')):
                os.mkdir(''.join((os.path.join(__dir__, '../'), report_dir_name)))
            file_name = os.path.join(__dir__, './report/{}{}'.format(now, '.xlsx'))
            self._wb.save(file_name)
        else:
            self._wb.save(file_fp)
        self._wb.close()
