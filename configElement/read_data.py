# !/uer/bin/env python3
# coding=utf-8
import datetime
import os

try:
    from openpyxl import load_workbook
except ImportError:
    os.system("pip3 install openpyxl")
    from openpyxl import load_workbook

from base.log import logged


class ReadData(object):
    @logged
    def __init__(self, file_name="case_data.xlsx", sheet_name="用例", start_row='A7'):
        self._wb = load_workbook(file_name)
        self._sheet = self._wb[sheet_name]
        self._case = []
        self.start_row = start_row

    @logged
    def get_email_data(self):
        """取excel中配置的email数据"""
        smtp_server = self._sheet['B4'].value
        res_user = self._sheet['B5'].value
        send_user = self._sheet['D4'].value
        send_pwd = self._sheet['D5'].value
        self._wb.close()
        return smtp_server, res_user, send_user, send_pwd

    @logged
    def get_public_data(self):
        """去接口参数的公共部分"""
        app_secret = self._sheet['F3'].value
        app_key = self._sheet['H3'].value
        version = self._sheet['B2'].value
        token = self._sheet['F5'].value
        return app_secret, app_key, version, token

    @logged
    def get_url(self):
        url = self._sheet['D3'].value
        return url

    @logged
    def get_interface_data(self):
        """去所有接口，返回接口列表"""
        index_no = int(self.start_row[1]) - 1
        while index_no <= self._sheet.max_row - 1:
            case_list = [column[index_no].value for column in self._sheet.columns]
            self._case.append(tuple(case_list[1:]))
            index_no += 1
        self._wb.close()
        return self._case

    @logged
    def write_the_result_to_the_new_excel(self, result, differences=None, file_fp=None):
        """把结果写入到excel中的最后两列"""
        end_column = self._sheet.max_column - 1
        no = 0
        for i in range(int(self.start_row[1]), self._sheet.max_row + 1):
            if differences is not None:
                self._sheet.cell(row=i, column=end_column).value = result[no]
                self._sheet.cell(row=i, column=self._sheet.max_column).value = differences[no]
                no += 1
        if file_fp is None:
            now = datetime.datetime.now().strftime('%Y-%m-%d_%H_%M_%S')
            file_name = os.path.join(os.path.dirname(os.path.realpath(__file__)), './result/result_') + now + '.xlsx'
            self._wb.save(file_name)
        else:
            self._wb.save(file_fp)
        self._wb.close()


if __name__ == '__main__':
    run = ReadData()
    case = run.get_interface_data()
    print(case)
    # run.write_the_result_to_the_new_excel('True')
